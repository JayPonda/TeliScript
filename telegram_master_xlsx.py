# telegram_master_xlsx.py - WITH TRANSLATION COLUMN
import os
from datetime import datetime
import hashlib
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from translator import AutoTranslator


class TelegramMasterXLSX:

    def __init__(self, master_file="data/telegram_messages_master.xlsx"):
        self.master_file = master_file
        self.existing_hashes = set()
        self.existing_ids = {}
        self.translator = AutoTranslator()

        # Ensure the data directory exists
        self._ensure_data_directory()

        # Load existing data
        self._load_existing_data()

    def _ensure_data_directory(self):
        """Ensure the data directory exists"""
        directory = os.path.dirname(self.master_file)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
            print(f"📁 Created data directory: {directory}")

    def _load_existing_data(self):
        """Load existing message hashes"""
        if os.path.exists(self.master_file):
            try:
                workbook = load_workbook(self.master_file)
                # Load from master sheet
                if "Master" in workbook.sheetnames:
                    master_sheet = workbook["Master"]
                    headers = [cell.value for cell in master_sheet[1]]

                    if headers:
                        # Find index of Message_Hash column
                        hash_index = None
                        channel_id_index = None
                        message_id_index = None

                        for i, header in enumerate(headers):
                            if header == "Message_Hash":
                                hash_index = i
                            elif header == "Channel_ID":
                                channel_id_index = i
                            elif header == "Message_ID":
                                message_id_index = i

                        # Process rows starting from row 2 (row 1 is header)
                        for row in master_sheet.iter_rows(min_row=2, values_only=True):
                            # Load message hash
                            if hash_index is not None and row[hash_index]:
                                self.existing_hashes.add(row[hash_index])

                            # Track by channel_id + message_id
                            if channel_id_index is not None and message_id_index is not None:
                                channel_id = row[channel_id_index] if row[channel_id_index] else ""
                                message_id = row[message_id_index] if row[message_id_index] else ""

                                if channel_id and message_id:
                                    if channel_id not in self.existing_ids:
                                        self.existing_ids[channel_id] = set()
                                    self.existing_ids[channel_id].add(message_id)

                        print(
                            f"📁 Loaded {len(self.existing_hashes)} existing messages"
                        )
            except Exception as e:
                print(f"❌ Error loading master file: {e}")
        else:
            print(f"📄 Creating new master file")

    def _generate_message_hash(self, message_data):
        """Generate unique hash for a message"""
        unique_string = f"{message_data['channel_id']}|{message_data['message_id']}|{message_data.get('datetime_utc', '')}"
        return hashlib.md5(unique_string.encode()).hexdigest()

    def _is_duplicate(self, message_data):
        """Check if message already exists"""
        # Check by hash
        msg_hash = self._generate_message_hash(message_data)
        if msg_hash in self.existing_hashes:
            return True

        # Check by channel_id + message_id
        channel_id = str(message_data.get("channel_id", ""))
        msg_id = str(message_data.get("message_id", ""))

        if channel_id in self.existing_ids and msg_id in self.existing_ids[channel_id]:
            return True

        return False

    def _clean_text(self, text):
        """Clean text for XLSX"""
        if not text:
            return ""
        text = str(text).replace("\0", "").replace("\x00", "").strip()
        return text[:4000]  # Limit length

    def _extract_links(self, text):
        """Extract markdown links from text"""
        if not text or not isinstance(text, str):
            return ""

        # Pattern to match markdown links: [](url)
        pattern = r'\]\(([^)]+)\)'
        matches = re.findall(pattern, text)

        # Return as comma-separated string
        return ",".join(matches)

    def add_messages(self, messages, channel_name):
        """Add messages with automatic translation"""
        if not messages:
            print(f"   📭 No messages to add")
            return 0

        new_messages = []
        duplicates = 0
        translated_count = 0

        for msg in messages:
            # Skip duplicates
            if self._is_duplicate(msg):
                duplicates += 1
                continue

            # Generate hash
            msg_hash = self._generate_message_hash(msg)
            msg["message_hash"] = msg_hash

            # Get original text
            original_text = self._clean_text(msg.get("text", ""))

            # Automatically translate any non-English text
            translated_text = self.translator.translate_text(original_text)

            if translated_text != original_text:
                translated_count += 1

            # Add both texts to message data
            msg["text_translated"] = translated_text

            # Extract links from original text
            links = self._extract_links(original_text)
            msg["links"] = links

            # Track in memory
            self.existing_hashes.add(msg_hash)
            channel_id = str(msg.get("channel_id", ""))
            msg_id = str(msg.get("message_id", ""))
            if channel_id and msg_id:
                if channel_id not in self.existing_ids:
                    self.existing_ids[channel_id] = set()
                self.existing_ids[channel_id].add(msg_id)

            new_messages.append(msg)

        if new_messages:
            # Create workbook if it doesn't exist
            if os.path.exists(self.master_file):
                workbook = load_workbook(self.master_file)
            else:
                workbook = Workbook()

                # Remove default sheet if it exists
                if "Sheet" in workbook.sheetnames:
                    workbook.remove(workbook["Sheet"])

            # Ensure Master sheet exists
            if "Master" not in workbook.sheetnames:
                master_sheet = workbook.create_sheet("Master")
            else:
                master_sheet = workbook["Master"]

            # Create channel sheet if it doesn't exist
            if channel_name not in workbook.sheetnames:
                channel_sheet = workbook.create_sheet(channel_name)
            else:
                channel_sheet = workbook[channel_name]

            # Define our headers
            our_headers = [
                "Date_Local_Date",
                "Channel_Name",
                "Translated_Text",
                "Links",
                "Message_Type",
                "Message_Hash",
                "Channel_ID",
                "Message_ID",
                "Global_ID",
                "Date_UTC",
                "Date_Local",
                "Sender_ID",
                "Sender_Name",
                "Views",
                "Forwards",
                "Added_At",
            ]

            # If Master sheet is empty, add headers
            if master_sheet.max_row == 1 and master_sheet.max_column == 1 and master_sheet.cell(1, 1).value is None:
                for col_idx, header in enumerate(our_headers, 1):
                    master_sheet.cell(row=1, column=col_idx, value=header)

            # If channel sheet is empty, add headers
            if channel_sheet.max_row == 1 and channel_sheet.max_column == 1 and channel_sheet.cell(1, 1).value is None:
                for col_idx, header in enumerate(our_headers, 1):
                    channel_sheet.cell(row=1, column=col_idx, value=header)

            # Get existing headers from Master sheet
            existing_master_headers = []
            if master_sheet.max_row > 0:
                existing_master_headers = [cell.value for cell in master_sheet[1]]

            # Get existing headers from Channel sheet
            existing_channel_headers = []
            if channel_sheet.max_row > 0:
                existing_channel_headers = [cell.value for cell in channel_sheet[1]]

            # Add new messages to sheets
            for msg in new_messages:
                # Prepare row data
                # Extract date part from Date_Local
                date_local_value = (
                    msg.get("datetime_local", "").strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                    if hasattr(msg.get("datetime_local", ""), "strftime")
                    else msg.get("datetime_local", "")
                )

                # Extract just the date part (YYYY-MM-DD)
                date_local_date_value = ""
                if date_local_value and " " in date_local_value:
                    date_local_date_value = date_local_value.split(" ")[0]
                elif date_local_value:
                    # If it's already just a date
                    date_local_date_value = date_local_value

                row_data = [
                    date_local_date_value,
                    channel_name,
                    msg.get("text_translated", ""),
                    msg.get("links", ""),
                    msg.get("media_type", "text"),
                    msg.get("message_hash", ""),
                    msg.get("channel_id", ""),
                    msg.get("message_id", ""),
                    msg.get("global_id", ""),
                    (
                        msg.get("datetime_utc", "").strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        if hasattr(msg.get("datetime_utc", ""), "strftime")
                        else msg.get("datetime_utc", "")
                    ),
                    date_local_value,
                    msg.get("sender_id", ""),
                    msg.get("sender_name", ""),
                    msg.get("views", 0),
                    msg.get("forwards", 0),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ]

                # Add to Master sheet
                master_sheet.append(row_data)

                # Add to Channel sheet
                channel_sheet.append(row_data)

            # Save workbook
            workbook.save(self.master_file)

            print(f"   ✅ Added {len(new_messages)} new messages")
            if translated_count > 0:
                print(f"   🔤 Translated {translated_count} non-English messages")
            if duplicates > 0:
                print(f"   ⏭️  Skipped {duplicates} duplicates")
        else:
            print(f"   ⏭️  All {len(messages)} messages already exist")

        return len(new_messages)

    def get_stats(self):
        """Get statistics"""
        stats = {
            "total_messages": len(self.existing_hashes),
            "channels": len(self.existing_ids),
            "file_size": (
                os.path.getsize(self.master_file)
                if os.path.exists(self.master_file)
                else 0
            ),
        }
        return stats
